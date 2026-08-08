"""
数据提取、转换和 Excel 输出
将粉笔 API 返回的原始数据转换为 Power BI 可用的结构化表格
"""
import os
import logging
from datetime import datetime, timezone

logger = logging.getLogger(__name__)

# Excel 表头定义（中文，与 Power BI 连接后可直接使用）
HEADERS = [
    "题目编号",
    "题目链接",
    "所属模块",
    "子模块",
    "题目类型",
    "练习时间",
    "用户答案",
    "正确答案",
    "是否正确",
    "题目来源",
    "题目年份",
    "考试类型",
    "难度",
    "试卷名称",
    "来源分类",
    "题库类型",
    "采集时间",
]

# 列宽配置
COLUMN_WIDTHS = {
    "A": 12, "B": 50, "C": 14, "D": 16, "E": 12,
    "F": 20, "G": 10, "H": 10, "I": 10, "J": 14,
    "K": 10, "L": 12, "M": 8, "N": 30, "O": 14,
    "P": 8, "Q": 20,
}


def transform_timestamp(ts) -> str:
    """将时间戳（毫秒或秒）转换为可读的日期时间字符串"""
    if not ts:
        return ""
    try:
        if isinstance(ts, (int, float)):
            # 粉笔 API 返回的时间戳通常是毫秒
            if ts > 1e12:
                ts = ts / 1000
            dt = datetime.fromtimestamp(ts, tz=timezone.utc)
            # 转换为北京时间
            from datetime import timedelta
            dt_bj = dt + timedelta(hours=8)
            return dt_bj.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(ts, str):
            # 尝试直接解析
            return ts.replace("T", " ").split(".")[0]
    except Exception as e:
        logger.debug("时间戳转换失败: %s (%s)", ts, e)
    return str(ts) if ts else ""


def normalize_subject(subject: str) -> str:
    """将题库类型代码转换为中文名称"""
    mapping = {
        "xingce": "行测",
        "shenlun": "申论",
    }
    return mapping.get(subject, subject)


def normalize_correct(is_correct: bool) -> str:
    """将布尔值转换为中文"""
    return "正确" if is_correct else "错误"


def transform_records(raw_records: list) -> list:
    """
    将原始 API 记录转换为标准化的表格行
    确保所有字段都有合适的格式和中文标签
    """
    transformed = []
    collect_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for record in raw_records:
        row = {
            "题目编号": record.get("question_id", ""),
            "题目链接": record.get("question_link", ""),
            "所属模块": record.get("module", ""),
            "子模块": record.get("sub_module", ""),
            "题目类型": record.get("question_type", ""),
            "练习时间": transform_timestamp(record.get("practice_time")),
            "用户答案": record.get("user_answer", ""),
            "正确答案": record.get("correct_answer", ""),
            "是否正确": normalize_correct(record.get("is_correct", False)),
            "题目来源": record.get("source", "") or record.get("source_label", ""),
            "题目年份": str(record.get("year", "")),
            "考试类型": record.get("exam_type", ""),
            "难度": str(record.get("difficulty", "")),
            "试卷名称": record.get("paper_name", ""),
            "来源分类": record.get("source_label", ""),
            "题库类型": normalize_subject(record.get("subject", "")),
            "采集时间": collect_time,
        }
        transformed.append(row)

    return transformed


def deduplicate_records(records: list) -> list:
    """
    去重：同一题目在同一练习中只保留一条记录
    去重键: 题目编号 + 试卷名称
    """
    seen = set()
    unique = []
    for r in records:
        key = (r.get("题目编号"), r.get("试卷名称"))
        if key not in seen:
            seen.add(key)
            unique.append(r)
    return unique


def save_to_csv(records: list, csv_path: str):
    """
    将练习记录保存为 CSV 文件（UTF-8 BOM 编码）
    Power BI Web 连接器可直接读取此文件

    每次运行完全覆盖旧文件，确保数据与粉笔最新状态一致
    """
    import csv

    os.makedirs(os.path.dirname(csv_path) or ".", exist_ok=True)

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=HEADERS)
        writer.writeheader()
        for record in records:
            writer.writerow(record)

    logger.info("CSV 文件已保存: %s (%d 条记录)", csv_path, len(records))


def save_to_excel(records: list, excel_path: str):
    """
    将练习记录保存到 Excel 文件
    如果文件已存在则追加/更新数据（按题目编号+试卷名称去重）
    """
    # 确保目录存在
    os.makedirs(os.path.dirname(excel_path) or ".", exist_ok=True)

    # 读取已有数据（如果文件存在）
    existing_records = []
    if os.path.exists(excel_path):
        try:
            existing_records = read_from_excel(excel_path)
            logger.info("读取已有数据 %d 条", len(existing_records))
        except Exception as e:
            logger.warning("读取已有Excel失败，将创建新文件: %s", e)

    # 合并新旧数据
    all_records = existing_records + records
    all_records = deduplicate_records(all_records)
    logger.info("合并去重后共 %d 条记录", len(all_records))

    # 写入 Excel
    write_to_excel(all_records, excel_path)
    logger.info("数据已保存到: %s", excel_path)


def write_to_excel(records: list, excel_path: str):
    """写入数据到 Excel 文件（带格式化）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "练习记录"

    # 写表头
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 写数据
    for row_idx, record in enumerate(records, 2):
        for col_idx, header in enumerate(HEADERS, 1):
            value = record.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            # 对错高亮
            if header == "是否正确":
                if value == "正确":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100")
                elif value == "错误":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006")

            # 题目链接设为超链接
            if header == "题目链接" and value:
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")

    # 设置列宽
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动筛选
    ws.auto_filter.ref = ws.dimensions

    wb.save(excel_path)


def read_from_excel(excel_path: str) -> list:
    """从 Excel 文件读取已有数据"""
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active

    records = []
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(h) if h else "" for h in row]
        break

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        record = {}
        for idx, header in enumerate(headers):
            if idx < len(row):
                record[header] = str(row[idx]) if row[idx] is not None else ""
        records.append(record)

    wb.close()
    return records
